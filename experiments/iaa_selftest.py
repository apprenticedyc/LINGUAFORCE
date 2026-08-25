# -*- coding: utf-8 -*-
"""Self-test: fill two templates with the SAME LLM labels; kappa must be 1.0."""
import json, os
from openpyxl import Workbook, load_workbook
from copy import copy

SAMPLE = r'E:\PythonCode\Paper\experiments\data\iaa\iaa_sample_150.json'
TMPL = r'E:\PythonCode\Paper\experiments\data\iaa\annotator_A.xlsx'
TA = r'E:\PythonCode\Paper\experiments\data\iaa\_test_a.xlsx'
TB = r'E:\PythonCode\Paper\experiments\data\iaa\_test_b.xlsx'
DIMS = ['D1','D2','D3','D4','D5','D6','D7']

rows = [json.loads(l) for l in open(SAMPLE, encoding='utf-8') if l.strip()]
vals = {}
for r in rows:
    d = r['dialogue_id']
    v = [r['gold_binary'], round(float(r['intensity']))]
    for dim in DIMS:
        v.append(int(r['dims'][dim]['level']))
    vals[d] = v

for path in (TA, TB):
    wb = load_workbook(TMPL)
    ws = wb['标注表']
    for row in ws.iter_rows(min_row=2):
        did = row[0].value
        if did is None:
            continue
        did = int(did)
        for j, x in enumerate(vals[did]):
            row[2 + j].value = x
    wb.save(path)
print('test files written')