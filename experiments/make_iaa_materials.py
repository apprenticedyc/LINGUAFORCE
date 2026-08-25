# -*- coding: utf-8 -*-
"""Build human IAA materials: stratified sample + Excel templates + guidelines.
Sample: 150 dialogues (75 coercive / 75 non-coercive) from the full release.
Two independent annotators each fill one Excel file; gold labels are NOT
included in the templates to avoid biasing annotators.
"""
import json, os, random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = r'E:\PythonCode\Paper'
FULL = os.path.join(ROOT, 'linguistic_agency_paper', 'data', 'linguaforce_full.jsonl')
OUT = os.path.join(ROOT, 'experiments', 'data', 'iaa')
os.makedirs(OUT, exist_ok=True)

random.seed(2026)
rows = [json.loads(l) for l in open(FULL, encoding='utf-8') if l.strip()]
pos = [r for r in rows if r['gold_binary'] == 1]
neg = [r for r in rows if r['gold_binary'] == 0]
N = 75
sample = random.sample(pos, N) + random.sample(neg, N)
random.shuffle(sample)
print('sample n =', len(sample), 'pos =', sum(1 for r in sample if r['gold_binary'] == 1))

# save sample (with gold, kept out of the annotator templates)
with open(os.path.join(OUT, 'iaa_sample_150.json'), 'w', encoding='utf-8') as f:
    for r in sample:
        f.write(json.dumps(r, ensure_ascii=False) + '\n')

DIMS = ['D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7']
DIM_NAME = {
    'D1': 'Directive Force (命令/指令强度)',
    'D2': 'Option Constraint (选项收窄/胁迫)',
    'D3': 'Normative Pressure (道德/规范施压)',
    'D4': 'Emotional Pressure (情绪施压)',
    'D5': 'Deceptiveness (欺骗性)',
    'D6': 'Toxicity (毒性/攻击性)',
    'D7': 'Explicitness (直白度)',
}
LEVEL_DESC = '0=None  1=Low  2=Moderate  3=High'

def make_template(path, sample):
    wb = Workbook()
    guide = wb.active
    guide.title = '标注说明'
    guide.column_dimensions['A'].width = 100
    lines = [
        ('人工标注说明（IAA）', True),
        ('', False),
        ('任务：阅读每条对话（双人轮流发言），然后判断是否含有“操纵性语言压力”并给出强度与七个维度的等级。', False),
        ('', False),
        ('第一步 是否存在操纵压力 (binary)：0=无，1=有', False),
        ('  有 = 说话人通过命令、胁迫、道德绑架、情绪勒索、欺骗、贬损等方式，推动或限制听话人的行为/选择。', False),
        ('', False),
        ('第二步 整体强度 (intensity)：0(无) 到 5(极强)', False),
        ('', False),
        ('第三步 七个维度等级 (D1-D7，每项填 0/1/2/3)：', False),
        ('  D1 命令强度：语气越接近命令/请求行动，分越高', False),
        ('  D2 选项收窄：威胁、最后通牒、无退路式表达', False),
        ('  D3 规范施压：道德义务、责任、身份绑架（如“你应该/你必须/不像话”）', False),
        ('  D4 情绪施压：内疚、羞耻、恐惧、卖惨作为筹码', False),
        ('  D5 欺骗性：歪曲事实、记忆否定、虚假信息', False),
        ('  D6 毒性：侮辱、谩骂、贬低、敌意', False),
        ('  D7 直白度：直说 vs 暗示/拐弯抹角', False),
        ('', False),
        ('填写示例：binary=1, intensity=3, D1=2, D2=1, D3=3, D4=2, D5=0, D6=0, D7=2', False),
        ('', False),
        ('注意：只填“标注表”sheet 中 E-O 列（红字列）；对话文本列请勿修改。', False),
        ('每条请独立判断，不要回看之前条目的答案。', False),
    ]
    for r, (t, bold) in enumerate(lines, 1):
        c = guide.cell(row=r, column=1, value=t)
        if bold:
            c.font = Font(bold=True, size=12)
    guide.freeze_panes = 'A2'

    ws = wb.create_sheet('标注表')
    headers = ['dialogue_id', '对话文本（只读）', 'binary (0/1)', 'intensity (0-5)'] + ['%s (%s)' % (d, DIM_NAME[d]) for d in DIMS]
    header_fill = PatternFill('solid', fgColor='DDEBF7')
    thin = Side(style='thin', color='BBBBBB')
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical='center')
        c.border = Border(bottom=thin)
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 90
    for d in DIMS:
        ws.column_dimensions[ws.cell(row=1, column=4 + DIMS.index(d)).column_letter].width = 30
    for i, r in enumerate(sample, 2):
        utts = '\n'.join('%d. %s' % (k + 1, u) for k, u in enumerate(r['utterances']))
        ws.cell(row=i, column=1, value=r['dialogue_id'])
        ws.cell(row=i, column=2, value=utts).alignment = Alignment(wrap_text=True, vertical='top')
        for col in range(3, 3 + 9):
            c = ws.cell(row=i, column=col)
            c.font = Font(color='C00000')
            c.border = Border(left=thin, right=thin)
    ws.freeze_panes = 'C2'
    wb.save(path)
    print('saved', path)

make_template(os.path.join(OUT, 'annotator_A.xlsx'), sample)
make_template(os.path.join(OUT, 'annotator_B.xlsx'), sample)
print('templates done')