# -*- coding: utf-8 -*-
"""Generate the human spot-check annotation template + reference file for LINGUAFORCE.

Stratified sample from the first-release (held-out) set: 120 items, balanced
60 pos / 60 neg by the corpus's binary coercion label (used only for coverage;
the annotator is not asked about coercion). Annotator fills:
  D: overall agency strength 0-5
  E: primary agency dimension D1-D7
  F: noticeable pressure present? 0/1
Reference file keeps the LLM 7-dim scores so agreement is computed vs the NEW
dimension annotations (not vs the old coercion labels).
"""
import json, os, random
from collections import defaultdict

SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                   "linguistic_agency_paper", "data", "linguaforce_first_release.jsonl")
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                       "linguistic_agency_paper", "data", "spotcheck")

SEED = 42
N_PER_CLASS = 60
SECOND_ANNOTATOR_N = 40
DIMS = ["D1", "D2", "D3", "D4", "D5", "D6", "D7"]

def pick_from_group(group, n, rng):
    buckets = defaultdict(list)
    for r in group:
        buckets[r["intensity"]].append(r)
    picked = []
    items = list(buckets.items())
    guard = 0
    while len(picked) < n and guard < n * 50:
        guard += 1
        rng.shuffle(items)
        for _k, lst in items:
            if lst and len(picked) < n:
                picked.append(lst.pop(rng.randrange(len(lst))))
    return picked

def main():
    rows = [json.loads(l) for l in open(SRC, encoding="utf-8")]
    rng = random.Random(SEED)
    pos = [r for r in rows if r["gold_binary"] == 1]
    neg = [r for r in rows if r["gold_binary"] == 0]
    sample = pick_from_group(pos, N_PER_CLASS, rng) + pick_from_group(neg, N_PER_CLASS, rng)
    rng.shuffle(sample)
    sample.sort(key=lambda r: r["dialogue_id"])
    second_ids = set(rng.sample([r["dialogue_id"] for r in sample], SECOND_ANNOTATOR_N))

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "spotcheck"
    headers = ["序号", "dialogue_id", "对话内容（A/B 轮流发言）",
               "整体话语能动性强度(0-5)", "最主要的能动性维度(D1-D7 选一个)",
               "是否存在明显施压?(0=无明显,1=明显)", "第二人标注?", "备注"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for i, r in enumerate(sample, 1):
        lines = []
        for j, u in enumerate(r["utterances"]):
            speaker = "A" if j % 2 == 0 else "B"
            lines.append(f"{speaker}: {u}")
        ws.append([i, r["dialogue_id"], "\n".join(lines), "", "", "",
                   "是" if r["dialogue_id"] in second_ids else "否", ""])

    widths = [6, 12, 90, 16, 22, 20, 12, 16]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            if cell.column == 2:
                cell.alignment = Alignment(vertical="top")
    ws.freeze_panes = "A2"
    annot_path = os.path.join(OUT_DIR, "spotcheck_annotate.xlsx")
    wb.save(annot_path)

    ref_path = os.path.join(OUT_DIR, "spotcheck_reference.csv")
    with open(ref_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write("dialogue_id,gold_binary,gold_multi,intensity," +
                ",".join(f"{d}_score" for d in DIMS) + "," +
                ",".join(f"{d}_level" for d in DIMS) +
                ",llm_agg_agency,llm_argmax_dim,llm_presence\n")
        for r in sample:
            mx = max(r["dims"][d]["score"] for d in DIMS)
            argmax = max(DIMS, key=lambda d: r["dims"][d]["score"])
            presence = 1 if mx >= 0.5 else 0
            vals = [str(r["dialogue_id"]), str(r["gold_binary"]), str(r["gold_multi"]), str(r["intensity"])]
            vals += [f'{r["dims"][d]["score"]:.2f}' for d in DIMS]
            vals += [str(r["dims"][d]["level"]) for d in DIMS]
            vals += [f"{mx*5:.2f}", argmax, str(presence)]
            f.write(",".join(vals) + "\n")

    from collections import Counter
    print("total sampled:", len(sample))
    print("binary coverage (pos/neg):", Counter(r["gold_binary"] for r in sample))
    print("intensity:", dict(sorted(Counter(r["intensity"] for r in sample).items())))
    print("second-annotator subset:", len(second_ids))
    print("saved:", annot_path)
    print("saved:", ref_path)

if __name__ == "__main__":
    main()
