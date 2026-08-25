# -*- coding: utf-8 -*-
"""Compute inter-annotator agreement from the two filled Excel templates.

Usage:
    python compute_iaa.py [--a path_a.xlsx] [--b path_b.xlsx] [--sample iaa_sample_150.json]
    (defaults point to experiments/data/iaa/)

Outputs per-variable Cohen's kappa (unweighted for binary, quadratic-weighted
for ordinal intensity and D1-D7), plus human-vs-LLM agreement on the sample.
"""
import json, os, sys, argparse
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.abspath(__file__))
DEFAULT_A = os.path.join(ROOT, 'data', 'iaa', 'annotator_A.xlsx')
DEFAULT_B = os.path.join(ROOT, 'data', 'iaa', 'annotator_B.xlsx')
DEFAULT_S = os.path.join(ROOT, 'data', 'iaa', 'iaa_sample_150.json')
DIMS = ['D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7']
VARS = ['binary', 'intensity'] + DIMS
KAPPA_COLS = {  # (name, max_category, quadratic_weighted)
    'binary': ('binary (0/1)', 1, False),
    'intensity': ('intensity (0-5)', 5, True),
    'D1': ('D1 (Directive Force)', 3, True),
    'D2': ('D2 (Option Constraint)', 3, True),
    'D3': ('D3 (Normative Pressure)', 3, True),
    'D4': ('D4 (Emotional Pressure)', 3, True),
    'D5': ('D5 (Deceptiveness)', 3, True),
    'D6': ('D6 (Toxicity)', 3, True),
    'D7': ('D7 (Explicitness)', 3, True),
}

def load_xlsx(path):
    wb = load_workbook(path, data_only=True)
    ws = wb['标注表']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        did = int(r[0])
        vals = {}
        # columns: 1=dialogue_id,2=text,3=binary,4=intensity,5..11=D1-D7
        raw = list(r[2:11])
        for v, var in zip(raw, VARS):
            if v is None or v == '':
                vals[var] = None
            else:
                vals[var] = int(round(float(v)))
        out[did] = vals
    return out

def cohens_kappa(a, b, maxcat, weighted):
    n = len(a)
    if n == 0:
        return float('nan')
    table = [[0.0] * (maxcat + 1) for _ in range(maxcat + 1)]
    for x, y in zip(a, b):
        table[x][y] += 1
    p_o = sum(table[i][i] for i in range(maxcat + 1)) / n
    px = [sum(table[i][j] for j in range(maxcat + 1)) / n for i in range(maxcat + 1)]
    py = [sum(table[i][j] for i in range(maxcat + 1)) / n for j in range(maxcat + 1)]
    if not weighted:
        p_e = sum(px[i] * py[i] for i in range(maxcat + 1))
        return (p_o - p_e) / (1 - p_e) if p_e != 1 else float('nan')
    num = den = 0.0
    for i in range(maxcat + 1):
        for j in range(maxcat + 1):
            w = (i - j) ** 2
            o = table[i][j] / n
            e = px[i] * py[j]
            num += w * o
            den += w * e
    return 1.0 - num / den if den else float('nan')

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--a', default=DEFAULT_A)
    ap.add_argument('--b', default=DEFAULT_B)
    ap.add_argument('--sample', default=DEFAULT_S)
    args = ap.parse_args()

    A = load_xlsx(args.a)
    B = load_xlsx(args.b)
    common = sorted(set(A) & set(B))
    print('pairs matched:', len(common))
    sample = [json.loads(l) for l in open(args.sample, encoding='utf-8') if l.strip()]
    llm = {r['dialogue_id']: r for r in sample}

    print(f'\n{"var":<12}{"H-H kappa":>12}{"n":>6}{"H-LLM kappa":>14}{"n":>6}')
    print('-' * 55)
    results = {}
    for var in VARS:
        name, maxcat, weighted = KAPPA_COLS[var]
        pairs = [(A[d][var], B[d][var]) for d in common
                 if A[d][var] is not None and B[d][var] is not None]
        k = cohens_kappa([p[0] for p in pairs], [p[1] for p in pairs], maxcat, weighted)
        # human vs LLM
        hp = []
        for d in common:
            if A[d][var] is not None and d in llm:
                if var == 'binary':
                    g = llm[d]['gold_binary']
                elif var == 'intensity':
                    g = int(round(float(llm[d]['intensity'])))
                else:
                    g = int(llm[d]['dims'][var]['level'])
                hp.append((A[d][var], g))
        kl = cohens_kappa([p[0] for p in hp], [p[1] for p in hp], maxcat, weighted)
        results[var] = (k, len(pairs), kl, len(hp))
        print(f'{var:<12}{k:>12.3f}{len(pairs):>6}{kl:>14.3f}{len(hp):>6}')

    print('\nInterpretation: kappa >= 0.60 substantial; 0.40-0.60 moderate.')

if __name__ == '__main__':
    main()